"""导入 API 路由"""
import logging
import os
import tempfile
import threading
import uuid
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import Response
from pydantic import BaseModel
from core.utils.database import db
from core.utils.inputer import CSVParser
from core.service.search import SearchService
from core.api.deps import require_wechat
from core.api.exceptions import BizError
from core.api.error_codes import ErrorCode
from core.utils.event_bus import emit
from datetime import datetime

logger = logging.getLogger("api_inputer")

router = APIRouter(
    prefix="/api/inputer",
    tags=["inputer"],
    dependencies=[Depends(require_wechat)],
)

# 公共路由：不需要微信连接即可访问（模板下载、文件上传）
public_router = APIRouter(
    prefix="/api/inputer",
    tags=["inputer-public"],
)


class CSVImportRequest(BaseModel):
    file_path: str


class ExcelImportRequest(BaseModel):
    file_path: str


class TencentDocImportRequest(BaseModel):
    doc_url: str = ""
    client_id: str = ""
    access_token: str = ""
    openid: str = ""



@public_router.get("/csv/template")
def download_csv_template():
    """下载 CSV 导入模板"""
    template_content = (
        "author_name,search_type,search_value\n"
        "示例作者,pages,1\n"
        "示例作者2,date,2026-01-01\n"
    )
    return Response(
        content=template_content,
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=author_template.csv"
        }
    )


@public_router.get("/excel/template")
def download_excel_template():
    """下载 Excel 导入模板"""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "作者导入"

        # 表头样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        header_alignment = Alignment(horizontal="center")

        # 写入表头
        headers = ["author_name", "search_type", "search_value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 写入示例数据
        sample_data = [
            ["示例作者1", "pages", "1"],
            ["示例作者2", "date", "2026-01-01"],
            ["示例作者3", "all", "10"],
        ]
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15

        # 写入内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=excel_template.xlsx"
            }
        )
    except ImportError:
        return {"code": -1, "msg": "服务器未安装 openpyxl 库"}


@public_router.post("/upload-csv")
async def upload_csv(file: UploadFile = File(...)):
    """上传 CSV 文件到临时目录，返回 file_path"""
    try:
        if not file.filename.lower().endswith('.csv'):
            return {"code": -1, "msg": "仅支持 .csv 文件"}

        content = await file.read()
        if len(content) == 0:
            return {"code": -1, "msg": "文件为空"}

        temp_dir = Path(tempfile.gettempdir()) / "video_monitor_csv"
        temp_dir.mkdir(parents=True, exist_ok=True)

        unique_name = f"{uuid.uuid4().hex[:8]}_{file.filename}"
        temp_path = temp_dir / unique_name
        temp_path.write_bytes(content)

        logger.info(f"[upload_csv] 文件已保存: {temp_path}")
        return {"code": 0, "data": {"file_path": str(temp_path)}, "msg": ""}
    except Exception as e:
        logger.error(f"[upload_csv] 上传失败: {e}")
        return {"code": -1, "msg": str(e)}


@router.post("/csv/import")
def import_csv(request: CSVImportRequest):
    """触发 CSV 后台导入，进度通过 SSE 推送"""
    if not os.path.exists(request.file_path):
        return {"code": -1, "msg": f"文件不存在: {request.file_path}"}

    def _run():
        try:
            _do_csv_import(request.file_path)
        except Exception as e:
            logger.error(f"[import_csv] 后台导入失败: {e}")
            emit("import_progress", {"phase": "done", "success": 0, "fail": 0, "total": 0, "error": str(e), "import_type": "csv"})

    threading.Thread(target=_run, daemon=True).start()
    return {"code": 0, "msg": "导入已开始"}


def _do_csv_import(file_path: str):
    """CSV 导入实际执行逻辑（后台线程），通过 SSE 推送进度"""
    from core.utils.store import Author, AuthorVideo

    parser = CSVParser(file_path)
    result = parser.validate()

    if not result.get("valid"):
        emit("import_progress", {"phase": "done", "success": 0, "fail": 0, "total": 0, "error": "CSV 格式错误", "import_type": "csv"})
        return

    data = result.get("data", [])
    total = len(data)

    if total == 0:
        emit("import_progress", {"phase": "done", "success": 0, "fail": 0, "total": 0, "import_type": "csv"})
        return

    emit("import_progress", {"phase": "start", "total": total, "import_type": "csv"})

    service = SearchService()
    success_count = 0
    fail_count = 0

    for idx, item in enumerate(data):
        author_name = item.get("author_name", "")
        search_type = item.get("search_type")
        search_value = item.get("search_value")

        try:
            search_result = service.search_author(author_name)
            if search_result.get("code") != 0 or not search_result.get("data"):
                fail_count += 1
                continue

            author_data = search_result.get("data")
            source_author_id = author_data.get("source_author_id")

            if search_type == "date":
                videos_result = service.get_author_videos_before_date(source_author_id, search_value)
            else:
                pages = int(search_value) if search_type in ["pages", "all"] else 1
                videos_result = service.get_author_videos(source_author_id, pages=pages)

            if videos_result.get("code") != 0:
                fail_count += 1
                continue

            videos = videos_result.get("data", [])

            latest_publish_date = None
            if videos:
                valid_times = [v.get("create_time", "") for v in videos if v.get("create_time")]
                if valid_times:
                    latest_publish_date = max(valid_times)

            author_id = f"doc_author_{int(datetime.now().timestamp() * 1000)}"
            author = Author(
                id=author_id,
                source_author_id=author_data.get("source_author_id"),
                name=author_data.get("name"),
                tag=None,
                bio=author_data.get("bio", ""),
                avatar_url=author_data.get("avatar_url", ""),
                cover_img_url=author_data.get("cover_img_url", ""),
                created_at=datetime.now().isoformat(),
                updated_at=datetime.now().isoformat(),
                latest_publish_date=latest_publish_date,
            )

            if not db.create_author(author):
                fail_count += 1
                continue

            for video in videos:
                author_video = AuthorVideo(
                    video_id=video.get('video_id', ''),
                    author_id=author_id,
                    title=video.get("title", ""),
                    object_nonce_id=video.get("object_nonce_id", ""),
                    url=video.get("url", ""),
                    spec=video.get("spec", ""),
                    file_size=video.get("file_size", 0),
                    cover_url=video.get("cover_url", ""),
                    decode_key=video.get("decode_key", 0),
                    author_avatar=video.get("author_avatar", ""),
                    duration=video.get("duration", 0),
                    create_time=video.get("create_time", ""),
                    is_downloaded=0,
                    download_path="",
                    downloaded_at=None,
                    video_type=video.get("video_type", "short_video"),
                )
                db.create_author_video(author_video)

            success_count += 1
        except Exception as e:
            logger.error(f"[_do_csv_import] 处理作者 {author_name} 失败: {e}")
            fail_count += 1
        finally:
            emit("import_progress", {
                "phase": "processing",
                "total": total,
                "current": idx + 1,
                "name": author_name,
                "success": success_count,
                "fail": fail_count,
                "import_type": "csv",
            })

    emit("import_progress", {
        "phase": "done",
        "total": total,
        "success": success_count,
        "fail": fail_count,
        "import_type": "csv",
    })


@public_router.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    """上传 Excel 文件到临时目录，返回 file_path"""
    try:
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return {"code": -1, "msg": "仅支持 .xlsx 或 .xls 文件"}

        content = await file.read()
        if len(content) == 0:
            return {"code": -1, "msg": "文件为空"}

        temp_dir = Path(tempfile.gettempdir()) / "video_monitor_excel"
        temp_dir.mkdir(parents=True, exist_ok=True)

        unique_name = f"{uuid.uuid4().hex[:8]}_{file.filename}"
        temp_path = temp_dir / unique_name
        temp_path.write_bytes(content)

        logger.info(f"[upload_excel] 文件已保存: {temp_path}")
        return {"code": 0, "data": {"file_path": str(temp_path)}, "msg": ""}
    except Exception as e:
        logger.error(f"[upload_excel] 上传失败: {e}")
        return {"code": -1, "msg": str(e)}


@router.post("/excel/import")
def import_excel(request: ExcelImportRequest):
    """触发 Excel 后台导入，进度通过 SSE 推送"""
    if not os.path.exists(request.file_path):
        return {"code": -1, "msg": f"文件不存在: {request.file_path}"}

    def _run():
        try:
            from core.utils.excel_client import ExcelClient
            client = ExcelClient(request.file_path)
            client.import_from_excel_to_database_with_progress()
        except Exception as e:
            logger.error(f"[import_excel] 后台导入失败: {e}")
            emit("import_progress", {"phase": "done", "success": 0, "fail": 0, "total": 0, "error": str(e), "import_type": "excel"})

    threading.Thread(target=_run, daemon=True).start()
    return {"code": 0, "msg": "导入已开始"}


@router.post("/tencent-doc/import")
def import_tencent_doc(request: TencentDocImportRequest):
    """触发腾讯文档后台导入，进度通过 SSE 推送"""
    if not request.doc_url:
        return ErrorCode.to_dict(ErrorCode.DOC_URL_INVALID, detail="请提供腾讯文档 URL")

    if not request.client_id or not request.access_token or not request.openid:
        return ErrorCode.to_dict(ErrorCode.DOC_CREDENTIAL_MISSING)

    def _run():
        try:
            from core.utils.tencent_doc import TencentDocClient
            client = TencentDocClient(
                doc_url=request.doc_url,
                client_id=request.client_id,
                access_token=request.access_token,
                openid=request.openid,
            )
            client.import_from_doc_to_database_with_progress()
        except BizError as e:
            e.log()
            emit("import_progress", {
                "phase": "done", "success": 0, "fail": 0, "total": 0,
                "error": e.detail or e.message,
                "error_code": e.error_code,
                "severity": e.severity,
                "import_type": "tencent_doc",
            })
        except Exception as e:
            logger.error(f"[import_tencent_doc] 后台导入失败: {e}")
            emit("import_progress", {"phase": "done", "success": 0, "fail": 0, "total": 0, "error": str(e), "import_type": "tencent_doc"})

    threading.Thread(target=_run, daemon=True).start()
    return {"code": 0, "msg": "导入已开始"}