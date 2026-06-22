"""Excel 客户端 - 批量导入视频数据到 Excel 文件"""

import logging
import os
from pathlib import Path
from typing import Dict, List, Optional, Any

logger = logging.getLogger("excel_client")


class ExcelClient:
    """Excel 批量导入客户端

    用于读写 Excel 文件，格式与腾讯文档一致：
    - author_name: 作者名
    - search_type: 搜索类型 (pages/date/all)
    - search_value: 搜索值 (页数或日期 YYYY-MM-DD)
    """

    # 期望的表头字段
    REQUIRED_HEADERS = ["author_name", "search_type", "search_value"]

    # search_type 有效值
    VALID_SEARCH_TYPES = ["pages", "date", "all"]

    def __init__(self, file_path: str):
        """初始化 Excel 客户端

        Args:
            file_path: Excel 文件路径
        """
        self.file_path = file_path
        self._workbook = None
        self._sheet = None

        # 尝试导入 openpyxl
        try:
            import openpyxl
            self._openpyxl = openpyxl
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        # 加载或创建工作簿
        self._load_workbook()

    def _load_workbook(self):
        """加载或创建工作簿"""
        if os.path.exists(self.file_path):
            self._workbook = self._openpyxl.load_workbook(self.file_path)
        else:
            self._workbook = self._openpyxl.Workbook()
            # 删除默认创建的工作表
            if "Sheet" in self._workbook.sheetnames:
                self._workbook.remove(self._workbook["Sheet"])

        # 确保有工作表
        if len(self._workbook.sheetnames) == 0:
            self._workbook.create_sheet("Sheet1")

        self._sheet = self._workbook.active

    def get_sheets(self) -> List[str]:
        """获取所有工作表名称

        Returns:
            工作表名称列表
        """
        return self._workbook.sheetnames

    def get_headers(self) -> List[str]:
        """获取表头

        Returns:
            表头列表
        """
        if self._sheet.max_row == 0:
            return []

        headers = []
        for cell in self._sheet[1]:
            value = cell.value
            if value:
                headers.append(str(value).strip())
            else:
                break

        return headers

    def get_all_data(self) -> List[Dict]:
        """获取所有数据（不含表头）

        Returns:
            数据行列表，每行是一个字典
        """
        headers = self.get_headers()
        if not headers:
            return []

        data = []
        for row_idx in range(2, self._sheet.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, start=1):
                cell_value = self._sheet.cell(row=row_idx, column=col_idx).value
                row_data[header] = cell_value

            # 跳过空行
            if any(row_data.values()):
                data.append(row_data)

        return data

    def write_author_data(self, data: List[Dict]):
        """写入作者数据

        Args:
            data: 作者数据列表，每项包含 author_name, search_type, search_value
        """
        # 写入表头
        for col_idx, header in enumerate(self.REQUIRED_HEADERS, start=1):
            self._sheet.cell(row=1, column=col_idx, value=header)

        # 写入数据
        for row_idx, row_data in enumerate(data, start=2):
            self._sheet.cell(row=row_idx, column=1, value=row_data.get("author_name", ""))
            self._sheet.cell(row=row_idx, column=2, value=row_data.get("search_type", ""))
            self._sheet.cell(row=row_idx, column=3, value=str(row_data.get("search_value", "")))

        # 保存文件
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
        self._workbook.save(self.file_path)
        logger.info(f"[Excel] 写入 {len(data)} 条数据到 {self.file_path}")

    def validate_structure(self) -> Dict:
        """验证 Excel 表结构

        Returns:
            {
                "valid": bool,
                "errors": List[str]
            }
        """
        errors = []

        # 验证工作表数量（只能有1个）
        sheets = self.get_sheets()
        if len(sheets) != 1:
            errors.append(f"Excel 只能有1个工作表，当前有 {len(sheets)} 个")
            return {"valid": False, "errors": errors}

        # 验证表头
        headers = self.get_headers()
        if not headers:
            errors.append("Excel 没有表头，请先在第一行添加表头")
            return {"valid": False, "errors": errors}

        for h in self.REQUIRED_HEADERS:
            if h not in headers:
                errors.append(f"表头缺少字段: {h}")

        if errors:
            return {"valid": False, "errors": errors}

        # 验证数据行
        data = self.get_all_data()
        for idx, row in enumerate(data, start=2):
            author_name = row.get("author_name", "").strip() if row.get("author_name") else ""
            search_type = row.get("search_type", "").strip() if row.get("search_type") else ""
            search_value = row.get("search_value", "")

            if not author_name:
                errors.append(f"第 {idx} 行: author_name 不能为空")

            if search_type not in self.VALID_SEARCH_TYPES:
                errors.append(f"第 {idx} 行: search_type 必须是 {self.VALID_SEARCH_TYPES} 之一")

            if search_type == "pages" or search_type == "all":
                try:
                    value = int(search_value)
                    if value <= 0:
                        errors.append(f"第 {idx} 行: search_value 必须是正整数")
                except (ValueError, TypeError):
                    errors.append(f"第 {idx} 行: search_value 必须是数字")
            elif search_type == "date":
                from datetime import datetime
                from core.utils.inputer import CSVParser
                normalized = CSVParser._normalize_date(str(search_value))
                if not normalized:
                    errors.append(f"第 {idx} 行: search_value 必须是有效日期格式, 实际 '{search_value}'")

        return {"valid": len(errors) == 0, "errors": errors}

    def close(self):
        """关闭工作簿"""
        if self._workbook:
            self._workbook.close()
            self._workbook = None
            self._sheet = None

    def import_from_excel_to_database(self) -> dict:
        """从 Excel 文件读取作者数据并保存到数据库

        完整流程：
        1. 读取 Excel 数据
        2. 验证表结构
        3. 解析作者数据
        4. 通过 SearchService 搜索作者并获取视频（统一标准化处理）
        5. 保存到数据库（作者表 + 视频表）

        Returns:
            {"code": 0, "data": {"authors_saved": N, "videos_saved": M}, "msg": ""}
            或 {"code": -1, "msg": "错误信息"}
        """
        from datetime import datetime
        from core.utils.database import db
        from core.utils.store import Author, AuthorVideo
        from core.service.search import SearchService

        # 1. 验证表结构
        validation_result = self.validate_structure()
        if not validation_result.get("valid"):
            errors = validation_result.get("errors", [])
            return {"code": -1, "msg": f"表结构验证失败: {'; '.join(errors)}"}

        # 2. 解析作者数据
        rows = self.get_all_data()
        if not rows:
            return {"code": -1, "msg": "Excel 没有作者数据"}

        if len(rows) == 1:
            logger.info(f"[import_from_excel_to_database] 只有1条数据，跳过批量导入测试")

        service = SearchService()
        authors_saved = 0
        videos_saved = 0

        # 3. 遍历每个作者
        for row in rows:
            author_name = str(row.get("author_name", "")).strip()
            search_type = str(row.get("search_type", "")).strip()
            search_value = str(row.get("search_value", ""))

            if not author_name:
                logger.warning("[import_from_excel_to_database] 跳过空作者名")
                continue

            # 通过 SearchService 搜索作者
            search_result = service.search_author(author_name)
            if search_result.get("code") != 0 or not search_result.get("data"):
                logger.warning(f"[import_from_excel_to_database] 未找到作者: {author_name}")
                continue

            author_data = search_result["data"]
            source_author_id = author_data.get("source_author_id", "")

            # 根据 search_type 分发到对应的视频拉取方法
            if search_type == "date":
                videos_result = service.get_author_videos_before_date(
                    source_author_id, search_value
                )
            elif search_type == "all":
                pages = int(search_value) if search_value else 50
                videos_result = service.get_author_videos(
                    source_author_id, pages=pages
                )
            else:
                try:
                    pages = int(search_value) if search_value else 1
                except ValueError:
                    logger.warning(f"[import_from_excel_to_database] 无效页数: {search_value}")
                    continue
                videos_result = service.get_author_videos(
                    source_author_id, pages=pages
                )

            if videos_result.get("code") != 0:
                logger.warning(f"[import_from_excel_to_database] 获取视频失败: {author_name}")
                continue

            videos = videos_result.get("data", [])
            if not videos:
                logger.warning(f"[import_from_excel_to_database] 作者 {author_name} 没有视频")
                continue

            # 4. 保存到数据库
            now = datetime.now().isoformat()

            # 计算最新视频发布时间
            latest_publish_date = None
            if videos:
                valid_times = [v.get("create_time", "") for v in videos if v.get("create_time")]
                if valid_times:
                    latest_publish_date = max(valid_times)

            existing_author = db.get_author_by_source_id(source_author_id)
            if existing_author:
                author_id = existing_author.id
                logger.info(f"[import_from_excel_to_database] 作者已存在: {author_name}")
            else:
                author_id = f"doc_author_{int(datetime.now().timestamp() * 1000)}"
                author = Author(
                    id=author_id,
                    source_author_id=source_author_id,
                    name=author_data.get("name", author_name),
                    tag="",
                    bio=author_data.get("bio", ""),
                    avatar_url=author_data.get("avatar_url", ""),
                    cover_img_url=author_data.get("cover_img_url", ""),
                    created_at=now,
                    updated_at=now,
                    latest_publish_date=latest_publish_date,
                )
                if db.create_author(author):
                    authors_saved += 1
                else:
                    logger.warning(f"[import_from_excel_to_database] 作者创建失败: {author_name}")
                    continue

            # 保存视频（SearchService 已标准化字段名）
            for video in videos:
                video_id = video.get('video_id', '')
                author_video = AuthorVideo(
                    video_id=video_id,
                    author_id=author_id,
                    title=video.get("title", ""),
                    object_nonce_id=video.get("object_nonce_id", ""),
                    url=video.get("url", ""),
                    spec=video.get("spec", ""),
                    file_size=video.get("file_size", 0),
                    cover_url=video.get("cover_url", ""),
                    decode_key=video.get("decode_key", 0),
                    author_avatar=video.get("author_avatar", ""),
                    duration=video.get("duration", 0),
                    create_time=video.get("create_time", ""),
                    is_downloaded=0,
                    download_path="",
                    downloaded_at=None,
                    video_type=video.get("video_type", "short_video"),
                )
                if db.create_author_video(author_video):
                    videos_saved += 1

        self.close()
        return {
            "code": 0,
            "data": {
                "authors_saved": authors_saved,
                "videos_saved": videos_saved
            },
            "msg": ""
        }

    def import_from_excel_to_database_with_progress(self):
        """从 Excel 文件读取作者数据并保存到数据库（带 SSE 进度推送）

        与 import_from_excel_to_database 逻辑相同，但通过 SSE 推送进度。
        """
        from core.utils.event_bus import emit

        validation_result = self.validate_structure()
        if not validation_result.get("valid"):
            errors = validation_result.get("errors", [])
            emit("import_progress", {"phase": "done", "success": 0, "fail": 0, "total": 0, "error": f"表结构验证失败: {'; '.join(errors)}", "import_type": "excel"})
            return

        rows = self.get_all_data()
        total = len(rows)

        if total == 0:
            emit("import_progress", {"phase": "done", "success": 0, "fail": 0, "total": 0, "import_type": "excel"})
            return

        emit("import_progress", {"phase": "start", "total": total, "import_type": "excel"})

        from datetime import datetime
        from core.utils.database import db
        from core.utils.store import Author, AuthorVideo
        from core.service.search import SearchService

        service = SearchService()
        authors_saved = 0
        videos_saved = 0

        for idx, row in enumerate(rows):
            author_name = str(row.get("author_name", "")).strip()
            search_type = str(row.get("search_type", "")).strip()
            search_value = str(row.get("search_value", ""))

            try:
                if not author_name:
                    continue

                search_result = service.search_author(author_name)
                if search_result.get("code") != 0 or not search_result.get("data"):
                    continue

                author_data = search_result["data"]
                source_author_id = author_data.get("source_author_id", "")

                if search_type == "date":
                    videos_result = service.get_author_videos_before_date(source_author_id, search_value)
                elif search_type == "all":
                    pages = int(search_value) if search_value else 50
                    videos_result = service.get_author_videos(source_author_id, pages=pages)
                else:
                    try:
                        pages = int(search_value) if search_value else 1
                    except ValueError:
                        continue
                    videos_result = service.get_author_videos(source_author_id, pages=pages)

                if videos_result.get("code") != 0:
                    continue

                videos = videos_result.get("data", [])
                if not videos:
                    continue

                now = datetime.now().isoformat()

                latest_publish_date = None
                if videos:
                    valid_times = [v.get("create_time", "") for v in videos if v.get("create_time")]
                    if valid_times:
                        latest_publish_date = max(valid_times)

                existing_author = db.get_author_by_source_id(source_author_id)
                if existing_author:
                    author_id = existing_author.id
                else:
                    author_id = f"doc_author_{int(datetime.now().timestamp() * 1000)}"
                    author = Author(
                        id=author_id,
                        source_author_id=source_author_id,
                        name=author_data.get("name", author_name),
                        tag="",
                        bio=author_data.get("bio", ""),
                        avatar_url=author_data.get("avatar_url", ""),
                        cover_img_url=author_data.get("cover_img_url", ""),
                        created_at=now,
                        updated_at=now,
                        latest_publish_date=latest_publish_date,
                    )
                    if db.create_author(author):
                        authors_saved += 1
                    else:
                        continue

                for video in videos:
                    video_id = video.get('video_id', '')
                    author_video = AuthorVideo(
                        video_id=video_id,
                        author_id=author_id,
                        title=video.get("title", ""),
                        object_nonce_id=video.get("object_nonce_id", ""),
                        url=video.get("url", ""),
                        spec=video.get("spec", ""),
                        file_size=video.get("file_size", 0),
                        cover_url=video.get("cover_url", ""),
                        decode_key=video.get("decode_key", 0),
                        author_avatar=video.get("author_avatar", ""),
                        duration=video.get("duration", 0),
                        create_time=video.get("create_time", ""),
                        is_downloaded=0,
                        download_path="",
                        downloaded_at=None,
                        video_type=video.get("video_type", "short_video"),
                    )
                    if db.create_author_video(author_video):
                        videos_saved += 1

            except Exception as e:
                logger.error(f"[import_from_excel_to_database_with_progress] 处理 {author_name} 失败: {e}")
            finally:
                emit("import_progress", {
                    "phase": "processing",
                    "total": total,
                    "current": idx + 1,
                    "name": author_name,
                    "success": authors_saved,
                    "fail": (idx + 1) - authors_saved,
                    "import_type": "excel",
                })

        self.close()
        emit("import_progress", {
            "phase": "done",
            "total": total,
            "success": authors_saved,
            "fail": total - authors_saved,
            "import_type": "excel",
        })

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()