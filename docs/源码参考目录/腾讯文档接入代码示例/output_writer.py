"""
输出写入器 - 策略模式

支持多种输出目标：CSV、腾讯文档、飞书文档。
通过 OutputConfig 配置切换，工厂方法 create_writer() 创建具体写入器。
"""

import logging
import os
import uuid
from abc import ABC, abstractmethod
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional

from core.titler.config.models import OutputConfig

logger = logging.getLogger(__name__)


def _get_env(name: str, default: str = "") -> str:
    """获取环境变量"""
    return os.environ.get(name, default)


class OutputWriter(ABC):
    """输出写入器抽象基类"""

    @abstractmethod
    def write(self, data: Dict[str, Any]) -> str:
        """写入一条数据，返回目标路径/标识"""
        raise NotImplementedError

    @abstractmethod
    def write_batch(self, data_list: List[Dict[str, Any]]) -> str:
        """批量写入多条数据"""
        raise NotImplementedError

    def _normalize(self, csv_data: dict, state: dict) -> dict:
        """统一数据结构（从 state 提取/兼容旧格式）"""
        logger.debug(f"_normalize called with csv_data={csv_data}, state_keys={list(state.keys())}")

        # 生成 task_id
        task_id = self._generate_task_id()

        if not csv_data:
            video_path = state.get("video_path", "")
            result = {
                "task_id": task_id,
                "video_path": video_path,
                "video_name": Path(video_path).name,
                "best_title": state.get("title", ""),
                "created_at": datetime.now().isoformat()
            }
            logger.debug(f"_normalize returning legacy format: {result}")
            return result
        result = {
            "task_id": task_id,
            "video_path": csv_data.get("video_path", ""),
            "video_name": csv_data.get("video_name", Path(csv_data.get("video_path", "")).name),
            "layout": csv_data.get("layout", ""),
            "kill": csv_data.get("kill", ""),
            "best_title": csv_data.get("best_title", ""),
            "alternative1": csv_data.get("alternative1", ""),
            "alternative2": csv_data.get("alternative2", ""),
            "alternative3": csv_data.get("alternative3", ""),
            "alternative4": csv_data.get("alternative4", ""),
            "alternative5": csv_data.get("alternative5", ""),
            "created_at": datetime.now().isoformat()
        }
        logger.debug(f"_normalize returning full format: {result}")
        return result

    @staticmethod
    def _generate_task_id() -> str:
        """生成唯一的任务标识符

        Returns:
            格式: task_20260411_212600_a3b5e2f1
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_suffix = uuid.uuid4().hex[:8]
        return f"task_{timestamp}_{unique_suffix}"


class ExcelWriter(OutputWriter):
    """Excel 文件写入器 - 按日期 sheet 和 task_id 索引的增量写入（与腾讯文档逻辑一致）"""

    # 固定列顺序，确保字段顺序一致
    FIXED_COLUMNS = [
        "task_id", "video_path", "video_name", "layout", "kill",
        "best_title",
        "alternative1", "alternative2", "alternative3", "alternative4", "alternative5",
        "created_at"
    ]

    def __init__(self, excel_path: str, mode: str = "a"):
        """
        初始化 Excel 写入器

        Args:
            excel_path: Excel 文件路径（直接写入该文件）
            mode: 写入模式，默认 "a"（追加）
        """
        if not excel_path:
            raise ValueError("ExcelWriter: excel_path 不能为空")
        self.excel_path = excel_path
        self.mode = mode
        logger.debug(f"ExcelWriter initialized with excel_path={excel_path}, mode={mode}")

    def _get_sheet_title(self) -> str:
        """获取基于日期的 sheet 标题"""
        return datetime.now().strftime("%Y-%m-%d")

    def _load_workbook(self, excel_path: str):
        """加载或创建 Workbook"""
        os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)
        if os.path.exists(excel_path):
            from openpyxl import load_workbook
            wb = load_workbook(excel_path)
            logger.debug(f"[Excel] 加载已有文件: {excel_path}, sheets: {wb.sheetnames}")
        else:
            from openpyxl import Workbook
            wb = Workbook()
            logger.debug(f"[Excel] 创建新文件: {excel_path}")
        return wb

    def _get_or_create_sheet(self, wb, sheet_title: str):
        """获取或创建指定标题的 sheet"""
        if sheet_title in wb.sheetnames:
            ws = wb[sheet_title]
            logger.debug(f"[Excel] 使用已有 sheet: {sheet_title}")
        else:
            ws = wb.create_sheet(sheet_title)
            logger.debug(f"[Excel] 创建新 sheet: {sheet_title}")
        return ws

    def _read_existing_data(self, ws) -> tuple:
        """
        读取现有 sheet 数据

        Returns:
            (rows: List[Dict], headers: List[str]) 元组
        """
        if ws.max_row <= 1:
            return [], []

        headers = [cell.value for cell in ws[1]]
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, start=1):
                row_data[header] = ws.cell(row=row_idx, column=col_idx).value
            rows.append(row_data)
        return rows, headers

    def _get_all_headers(self, data_list: List[Dict], existing_headers: List[str]) -> List[str]:
        """获取所有可能的表头（合并现有和新的），保持固定顺序"""
        # 收集所有出现的列名
        all_columns = set(existing_headers)
        for data in data_list:
            all_columns.update(data.keys())

        # 按固定顺序排列，FIXED_COLUMNS 中没有的列按发现顺序添加在末尾
        ordered = []
        for col in self.FIXED_COLUMNS:
            if col in all_columns:
                ordered.append(col)
                all_columns.remove(col)

        # 剩余列按原有顺序（集合遍历顺序）
        ordered.extend(sorted(all_columns))

        return ordered

    def write(self, data: Dict[str, Any]) -> str:
        """写入单条数据（以 task_id 为索引，幂等写入）"""
        return self.write_batch([data])

    def write_batch(self, data_list: List[Dict[str, Any]]) -> str:
        """
        批量写入数据（以 task_id 为索引的增量写入）

        逻辑：
        1. 直接写入用户指定的 Excel 文件
        2. 在文件中按日期创建 sheet（如 "2026-04-17"）
        3. 以 task_id 为主键查找已存在的行
        4. 已存在则更新，不存在则追加
        """
        if not data_list:
            return self.excel_path

        excel_path = self.excel_path
        sheet_title = self._get_sheet_title()
        logger.info(f"[Excel] 批量写入 {len(data_list)} 条: {excel_path}, sheet: {sheet_title}")

        # 加载或创建 workbook
        wb = self._load_workbook(excel_path)

        # 获取或创建 sheet
        ws = self._get_or_create_sheet(wb, sheet_title)

        # 读取现有数据
        existing_rows, existing_headers = self._read_existing_data(ws)
        all_headers = self._get_all_headers(data_list, existing_headers)

        # 构建 task_id -> 行索引的映射（用于快速查找）
        task_id_to_row = {}
        for i, row in enumerate(existing_rows, start=2):  # 从第2行开始（第1行是表头）
            if row.get("task_id"):
                task_id_to_row[row["task_id"]] = i

        # 处理数据：更新已存在或追加新数据
        updated_count = 0
        added_count = 0
        for data in data_list:
            task_id = data.get("task_id", "")
            if task_id and task_id in task_id_to_row:
                # 更新已存在的行
                row_idx = task_id_to_row[task_id]
                for col_idx, header in enumerate(all_headers, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=data.get(header, ""))
                updated_count += 1
                logger.debug(f"[Excel] 更新已存在行: task_id={task_id}, row={row_idx}")
            else:
                # 追加新行
                row_idx = ws.max_row + 1
                for col_idx, header in enumerate(all_headers, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=data.get(header, ""))
                if task_id:
                    task_id_to_row[task_id] = row_idx
                added_count += 1
                logger.debug(f"[Excel] 追加新行: task_id={task_id}, row={row_idx}")

        # 写入表头（如果需要）
        if not existing_headers and all_headers:
            for col_idx, header in enumerate(all_headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

        # 保存文件
        wb.save(excel_path)
        logger.info(f"[Excel] 完成: 新增 {added_count} 行, 更新 {updated_count} 行, 总计 {len(existing_rows) + added_count} 行")
        return excel_path


class TencentDocWriter(OutputWriter):
    """腾讯文档写入器（REST API v3）"""

    BASE_URL = "https://docs.qq.com"

    def __init__(self, doc_url: str = ""):
        # 自动加载 .env 文件
        try:
            from dotenv import load_dotenv
            # 查找 .env 文件（从当前目录向上查找）
            current_path = Path.cwd()
            logger.debug(f"[TencentDoc] cwd={current_path}")
            for _ in range(5):  # 最多向上查找5层
                env_path = current_path / ".env"
                logger.debug(f"[TencentDoc] 检查 .env: {env_path} (exists={env_path.exists()})")
                if env_path.exists():
                    load_dotenv(env_path)
                    logger.info(f"[TencentDoc] 已加载 .env 文件: {env_path}")
                    break
                if current_path.parent == current_path:  # 到达根目录
                    break
                current_path = current_path.parent
        except ImportError:
            logger.warning("[TencentDoc] python-dotenv 未安装，无法自动加载 .env")

        # 如果 doc_url 未提供，从环境变量读取
        if not doc_url:
            doc_url = _get_env("TENCENT_DOC_URL", "")

            # 如果环境变量也没有完整 URL，尝试从 ID 和 Tab ID 构建
            if not doc_url:
                doc_id = _get_env("TENCENT_DOC_ID", "")
                tab_id = _get_env("TENCENT_DOC_TAB_ID", "")
                if doc_id:
                    doc_url = f"https://docs.qq.com/sheet/{doc_id}"
                    if tab_id:
                        doc_url += f"?tab={tab_id}"
                    logger.info(f"[TencentDoc] 从环境变量构建 URL: {doc_url}")

            # 如果仍然没有 URL，抛出错误
            if not doc_url:
                raise ValueError(
                    "TencentDocWriter: doc_url 不能为空，请在 .env 中配置 TENCENT_DOC_URL "
                    "或提供 TENCENT_DOC_ID 和 TENCENT_DOC_TAB_ID"
                )

        self.doc_url = doc_url
        self.file_id, self.sheet_id = self._parse_doc_url(doc_url)

        # 添加工作表缓存（避免重复API调用）
        self._sheets_cache: Dict[str, Dict[str, Any]] = {}

        logger.info(f"[TencentDoc] fileId={self.file_id}, sheetId={self.sheet_id or '(使用默认)'}")

        self.access_token = _get_env("TENCENT_DOC_ACCESS_TOKEN")
        self.client_id = _get_env("TENCENT_DOC_CLIENT_ID")
        self.openid = _get_env("TENCENT_DOC_OPENID")

        logger.debug(f"[TencentDoc] 凭证读取: access_token={'已设置' if self.access_token else '空'}, client_id={'已设置' if self.client_id else '空'}, openid={'已设置' if self.openid else '空'}")

        missing = []
        if not self.access_token:
            missing.append("TENCENT_DOC_ACCESS_TOKEN")
        if not self.client_id:
            missing.append("TENCENT_DOC_CLIENT_ID")
        if not self.openid:
            missing.append("TENCENT_DOC_OPENID")
        if missing:
            raise ValueError(
                f"TencentDocWriter: .env 缺少凭证: {', '.join(missing)}"
            )

        logger.info(f"TencentDocWriter: fileId={self.file_id}, sheetId={self.sheet_id or '(默认)'}")

    def _parse_doc_url(self, url: str) -> tuple:
        import re
        match = re.search(r'/sheet/([^/?#]+)', url)
        if not match:
            raise ValueError(f"无法从 URL 提取 fileId: {url}")
        file_id = match.group(1)
        tab_match = re.search(r'[?&]tab=([^&#]+)', url)
        sheet_id = tab_match.group(1) if tab_match else ""
        return file_id, sheet_id

    def _headers(self) -> dict:
        return {
            "Access-Token": self.access_token,
            "Client-Id": self.client_id,
            "Open-Id": self.openid,
            "Content-Type": "application/json",
        }

    def _request(self, method: str, path: str, **kwargs) -> dict:
        import requests
        url = f"{self.BASE_URL}{path}"
        kwargs.setdefault("timeout", 30)
        kwargs.setdefault("headers", self._headers())
        response = requests.request(method, url, **kwargs)
        response.raise_for_status()
        data = response.json()
        if isinstance(data, dict):
            if data.get("code", 0) != 0 and "ret" not in data:
                raise RuntimeError(f"API error: code={data.get('code')}, msg={data.get('message')}")
            if data.get("ret", 0) != 0:
                raise RuntimeError(f"API error: ret={data.get('ret')}, msg={data.get('msg')}")
        return data

    def _query_sheets(self, use_cache: bool = True) -> List[Dict[str, Any]]:
        """查询文档中的所有工作表

        Args:
            use_cache: 是否使用缓存（默认True）

        Returns:
            工作表列表，每个元素包含 title 和 sheetId
        """
        # 使用缓存（如果可用）
        if use_cache and self._sheets_cache:
            return list(self._sheets_cache.values())

        # 使用正确的 API 端点：/files/{file_id} 返回 properties 数组
        path = f"/openapi/spreadsheet/v3/files/{self.file_id}"
        data = self._request("GET", path)
        # properties 数组包含所有 sheet 信息
        sheets = data.get("properties", [])

        # 更新缓存
        self._sheets_cache.clear()
        for sheet in sheets:
            sheet_id = sheet.get("sheetId", "")
            if sheet_id:
                self._sheets_cache[sheet_id] = sheet

        return sheets

    def _find_sheet_by_title(self, title: str) -> Optional[str]:
        """根据标题查找工作表

        Args:
            title: 工作表标题

        Returns:
            找到的 sheetId，未找到返回 None
        """
        sheets = self._query_sheets()
        logger.info(f"[TencentDoc] 查找同名工作表: '{title}'，当前文档共有 {len(sheets)} 个工作表")
        for idx, sheet in enumerate(sheets):
            sheet_title = sheet.get("title", "")
            sheet_id = sheet.get("sheetId", "")
            logger.info(f"[TencentDoc]   工作表{idx + 1}: title='{sheet_title}', id='{sheet_id}'")
            if sheet_title == title:
                logger.info(f"[TencentDoc] 找到已存在的工作表: {title} (ID: {sheet_id})")
                return sheet_id
        logger.info(f"[TencentDoc] 未找到同名工作表: '{title}'")
        return None

    def _create_sheet(self, title: str) -> str:
        """创建新工作表

        Args:
            title: 工作表标题

        Returns:
            新创建的工作表 ID
        """
        body = {
            "requests": [{"addSheetRequest": {"title": title}}]
        }
        path = f"/openapi/spreadsheet/v3/files/{self.file_id}/batchUpdate"
        data = self._request("POST", path, json=body)
        responses = data.get("responses", [])
        if not responses:
            raise RuntimeError("创建工作表失败: API 未返回响应")

        sheet_id = responses[0].get("addSheetResponse", {}).get("properties", {}).get("sheetId", "")

        # 添加到缓存（新创建的工作表初始行数为0）
        self._sheets_cache[sheet_id] = {
            "sheetId": sheet_id,
            "title": title,
            "rowCount": 0,
            "columnCount": 0
        }

        return sheet_id

    def _get_sheet_data_row_count(self, sheet_id: str) -> int:
        """获取 sheet 中已有数据的行数

        Args:
            sheet_id: 工作表 ID

        Returns:
            数据行数 (不含表头)，至少返回1避免覆盖表头
        """
        # 优先从缓存获取
        if sheet_id in self._sheets_cache:
            sheet = self._sheets_cache[sheet_id]
            row_count = sheet.get("rowCount", 0)
            logger.info(f"[TencentDoc] 工作表 {sheet_id} 实际数据行数: {row_count} (从缓存)")
            # 安全返回：至少返回1，避免覆盖表头
            return max(1, row_count)

        # 缓存未命中，刷新缓存
        logger.info(f"[TencentDoc] 工作表 {sheet_id} 缓存未命中，刷新缓存")
        self._query_sheets(use_cache=False)

        # 再次尝试从缓存获取
        if sheet_id in self._sheets_cache:
            sheet = self._sheets_cache[sheet_id]
            row_count = sheet.get("rowCount", 0)
            logger.info(f"[TencentDoc] 工作表 {sheet_id} 实际数据行数: {row_count} (刷新后)")
            return max(1, row_count)

        # 仍然找不到，返回安全默认值1（从第2行开始，避免覆盖表头）
        logger.warning(f"[TencentDoc] 无法找到工作表 {sheet_id}，使用默认值 1")
        return 1

    def write(self, data: Dict[str, Any]) -> str:
        return self.write_batch([data])

    def write_batch(self, data_list: List[Dict[str, Any]]) -> str:
        """批量写入数据，支持追加模式

        Args:
            data_list: 数据列表

        Returns:
            文档 URL
        """
        logger.info(f"[TencentDoc] 批量写入 {len(data_list)} 行")

        # 使用年月日格式作为 sheet 标题
        sheet_title = datetime.now().strftime("%Y-%m-%d")

        # 如果没有 sheet_id，查询所有 sheets 找到第一个
        if not self.sheet_id:
            logger.info(f"[TencentDoc] 未指定 sheetId，查询默认 sheet")
            sheets = self._query_sheets()
            if sheets and len(sheets) > 0:
                self.sheet_id = sheets[0].get("sheetId", "")
                logger.info(f"[TencentDoc] 使用默认 sheet: {sheets[0].get('title', '')} (ID: {self.sheet_id})")
            else:
                logger.warning(f"[TencentDoc] 无法找到任何 sheet，将创建新 sheet")

        # 优先检查是否已存在同名 sheet（按标题匹配）
        # 这个检查应该在所有情况下执行，无论是否有初始 sheet_id
        existing_sheet_id = self._find_sheet_by_title(sheet_title)

        if existing_sheet_id:
            # 找到同名工作表，追加到现有 sheet
            sheet_id = existing_sheet_id
            is_new_sheet = False
            logger.info(f"[TencentDoc] 找到同名工作表，追加数据: {sheet_title} (ID: {sheet_id})")
        else:
            # 没有找到同名工作表，创建新 sheet
            sheet_id = self._create_sheet(sheet_title)
            is_new_sheet = True
            logger.info(f"[TencentDoc] 创建新工作表: {sheet_title} (ID: {sheet_id})")

        # 只在新 sheet 时写入表头
        if is_new_sheet and data_list:
            header_values = []
            for k in data_list[0].keys():
                header_values.append({"cellValue": {"text": str(k)}})
            header_body = {
                "requests": [{
                    "updateRangeRequest": {
                        "sheetId": sheet_id,
                        "gridData": {
                            "startRow": 0,
                            "startColumn": 0,
                            "rows": [{"values": header_values}],
                        }
                    }
                }]
            }
            path = f"/openapi/spreadsheet/v3/files/{self.file_id}/batchUpdate"
            self._request("POST", path, json=header_body)

        # 写数据行
        rows_data = []
        for data in data_list:
            values = []
            for v in data.values():
                values.append({"cellValue": {"text": str(v) if v is not None else ""}})
            rows_data.append({"values": values})

        # 追加模式: 从现有数据的最后一行之后开始写入
        # 新 sheet: 从第2行开始 (第1行是表头)
        if not is_new_sheet:
            # 查询现有 sheet 的数据量
            start_row = self._get_sheet_data_row_count(sheet_id)
        else:
            start_row = 1

        logger.info(f"[TencentDoc] 准备写入: sheetId={sheet_id}, startRow={start_row}, rows={len(rows_data)}")

        data_body = {
            "requests": [{
                "updateRangeRequest": {
                    "sheetId": sheet_id,
                    "gridData": {
                        "startRow": start_row,
                        "startColumn": 0,
                        "rows": rows_data,
                    }
                }
            }]
        }
        path = f"/openapi/spreadsheet/v3/files/{self.file_id}/batchUpdate"
        self._request("POST", path, json=data_body)
        logger.info(f"[TencentDoc] 批量写入完成: sheet={sheet_title}, rows={start_row}-{start_row + len(rows_data)}")
        return self.doc_url


class FeishuWriter(OutputWriter):
    """飞书文档写入器（预留）"""

    def __init__(self, app_id: str, table_id: str, app_secret: str = ""):
        if not app_id:
            raise ValueError("FeishuWriter: app_id 不能为空")
        if not table_id:
            raise ValueError("FeishuWriter: table_id 不能为空")
        self.app_id = app_id
        self.table_id = table_id
        self.app_secret = app_secret
        logger.debug(f"FeishuWriter initialized with app_id={app_id}, table_id={table_id}")

    def write(self, data: Dict[str, Any]) -> str:
        raise NotImplementedError("飞书文档写入待实现")

    def write_batch(self, data_list: List[Dict[str, Any]]) -> str:
        raise NotImplementedError("飞书文档批量写入待实现")


def create_writer(config: OutputConfig) -> OutputWriter:
    """根据配置创建写入器"""
    logger.debug(f"create_writer: output_type={config.output_type}")
    if config.output_type == "excel":
        writer = ExcelWriter(excel_path=config.xlsx_path)
    elif config.output_type == "tencent":
        writer = TencentDocWriter(doc_url=config.tencent_doc_url)
    elif config.output_type == "feishu":
        writer = FeishuWriter(
            app_id=config.feishu_app_id,
            table_id=config.feishu_table_id,
            app_secret=config.feishu_app_secret,
        )
    else:
        raise ValueError(f"未知的输出类型: {config.output_type}")
    logger.debug(f"create_writer: {type(writer).__name__}")
    return writer


def write_output_node(state: dict) -> dict:
    """通用输出节点 - 根据配置写入不同目标

    替代旧的 write_title_csv_node，通过 OutputConfig 选择写入器。
    """
    logger.debug(f"write_output_node called with state_keys={list(state.keys())}")

    # 获取配置（支持 dict 或 OutputConfig 对象）
    config_data = state.get("output_config")
    if config_data is None:
        from ..config.models import OutputConfig
        config = OutputConfig()
    elif isinstance(config_data, dict):
        from ..config.models import OutputConfig
        config = OutputConfig(**config_data)
    else:
        config = config_data

    csv_data = state.get("csv_data", {})

    writer = create_writer(config)
    data = writer._normalize(csv_data, state)
    path = writer.write(data)

    logger.info(f"[输出] 写入完成: {path}")
    if data.get("best_title"):
        logger.info(f"[输出] 最佳标题: {data.get('best_title', 'N/A')}")

    logger.debug(f"write_output_node returning csv_output_path={path}")
    return {"csv_output_path": path}


# DEPRECATED: 旧接口保留，向后兼容
def write_title_csv_node(state: dict) -> dict:
    """向后兼容的旧节点函数（已废弃）"""
    return write_output_node(state)
